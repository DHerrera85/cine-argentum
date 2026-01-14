import json
import re
import sys
from pathlib import Path
from typing import Dict, List, Any

try:
    from openpyxl import load_workbook
except ImportError:
    print("Missing dependency: openpyxl. Please install it (pip install openpyxl).", file=sys.stderr)
    sys.exit(1)

# Configurable header synonyms (case-insensitive)
HEADER_MAP = {
    'title': ['title', 'titulo', 'título', 'nombre', 'pelicula', 'película'],
    'year': ['year', 'anio', 'año'],
    'genre': ['genre', 'genero', 'género'],
    'actors': ['actors', 'actores', 'reparto', 'elenco'],
    'channel': ['channel', 'canal', 'plataforma', 'distribuidor'],
    'image': ['image', 'imagen', 'poster'],
    'orientation': ['orientation', 'orientacion', 'orientación'],
    'type': ['type', 'tipo'],
    'synopsis': ['synopsis', 'sinopsis', 'descripcion', 'descripción', 'resumen']
}

DEFAULTS = {
    'orientation': 'vertical',
    'type': 'pelicula'
}

ID_PREFIX = 'P'  # IDs for movies appended from Excel
ID_START = 1000   # start high to avoid collisions with existing ids


def normalize_header(name: str) -> str:
    n = name.strip().lower()
    for key, syns in HEADER_MAP.items():
        if n in syns:
            return key
    return n  # unknown; will be ignored unless matches schema


def parse_excel_rows(xlsx_path: Path) -> List[Dict[str, Any]]:
    wb = load_workbook(filename=str(xlsx_path), read_only=True, data_only=True)
    ws = wb.worksheets[0]  # first sheet
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(c or '').strip() for c in rows[0]]
    header_keys = [normalize_header(h) for h in headers]

    items: List[Dict[str, Any]] = []
    for r in rows[1:]:
        if not any(r):
            continue  # skip empty
        item: Dict[str, Any] = {}
        for idx, val in enumerate(r):
            key = header_keys[idx] if idx < len(header_keys) else None
            if not key or not val:
                continue
            if key == 'title':
                item['title'] = str(val).strip()
            elif key == 'year':
                item['year'] = str(val).strip()
            elif key == 'genre':
                item['genre'] = str(val).strip()
            elif key == 'actors':
                # split by comma / semicolon
                actors = [a.strip() for a in re.split(r'[;,]', str(val)) if a and a.strip()]
                item['actors'] = actors
            elif key == 'channel':
                item['channel'] = str(val).strip()
            elif key == 'image':
                item['image'] = str(val).strip()
            elif key == 'orientation':
                v = str(val).strip().lower()
                item['orientation'] = 'horizontal' if v.startswith('h') else 'vertical'
            elif key == 'type':
                v = str(val).strip().lower()
                item['type'] = 'pelicula' if 'pel' in v else v
            elif key == 'synopsis':
                item['synopsis'] = str(val).strip()
            # ignore unknown keys
        if 'title' not in item:
            continue  # minimal required
        # defaults
        item.setdefault('year', '')
        item.setdefault('genre', '')
        item.setdefault('actors', [])
        item.setdefault('channel', '')
        item.setdefault('image', '')
        item.setdefault('orientation', DEFAULTS['orientation'])
        item.setdefault('type', DEFAULTS['type'])
        # create subtitle from genre if available
        item.setdefault('subtitle', item.get('genre') or '')
        item.setdefault('synopsis', '')
        items.append(item)
    return items


def next_id(existing_ids: List[str]) -> str:
    max_num = ID_START - 1
    for eid in existing_ids:
        m = re.match(r'^[Pp](\d+)$', str(eid))
        if m:
            num = int(m.group(1))
            if num > max_num:
                max_num = num
    return f"{ID_PREFIX}{max_num + 1}"


def merge_items(existing: List[Dict[str, Any]], new_items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    # map existing by (title_lower, year)
    existing_map: Dict[tuple, Dict[str, Any]] = {}
    for it in existing:
        key = (str(it.get('title', '')).strip().lower(), str(it.get('year', '')).strip())
        existing_map[key] = it

    existing_ids = [str(it.get('id', '')) for it in existing]
    merged = list(existing)

    for new in new_items:
        key = (str(new.get('title', '')).strip().lower(), str(new.get('year', '')).strip())
        if key in existing_map:
            # update existing with synopsis from Excel if provided
            cur = existing_map[key]
            syn = str(new.get('synopsis', '')).strip()
            if syn:
                cur['synopsis'] = syn
            # ensure type pelicula for movies
            if not cur.get('type'):
                cur['type'] = 'pelicula'
            # optionally fill missing channel/image from Excel
            if (not str(cur.get('channel', '')).strip()) and str(new.get('channel', '')).strip():
                cur['channel'] = new['channel']
            if (not str(cur.get('image', '')).strip()) and str(new.get('image', '')).strip():
                cur['image'] = new['image']
            # subtitle from genre if missing
            if (not str(cur.get('subtitle', '')).strip()) and str(new.get('genre', '')).strip():
                cur['subtitle'] = new['genre']
            # keep orientation as is; do not overwrite
            continue
        # else: assign new id and append
        nid = next_id(existing_ids)
        new['id'] = nid
        existing_ids.append(nid)
        merged.append(new)
        existing_map[key] = new
    return merged


def main():
    # Args: excel_path [json_path]
    if len(sys.argv) < 2:
        print("Usage: merge_excel_to_datajson.py <excel_path> [data_json_path]", file=sys.stderr)
        sys.exit(2)
    excel_path = Path(sys.argv[1])
    json_path = Path(sys.argv[2]) if len(sys.argv) > 2 else Path(__file__).resolve().parent.parent / 'data.json'

    if not excel_path.exists():
        print(f"Excel not found: {excel_path}", file=sys.stderr)
        sys.exit(3)
    if not json_path.exists():
        print(f"JSON not found: {json_path}", file=sys.stderr)
        sys.exit(4)

    with json_path.open('r', encoding='utf-8') as f:
        data = json.load(f)
    items = data.get('items', [])

    new_items = parse_excel_rows(excel_path)
    if not new_items:
        print("No rows parsed from Excel.", file=sys.stderr)
        sys.exit(5)

    merged = merge_items(items, new_items)
    data['items'] = merged

    # Write back preserving utf-8; keep compact formatting similar to existing (no extra spaces)
    with json_path.open('w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, separators=(',', ': '))

    print(f"Merged {len(new_items)} new item(s). Total: {len(merged)}")


if __name__ == '__main__':
    main()
